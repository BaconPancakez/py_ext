import openpyxl

# Load the workbook and select the sheet
wb = openpyxl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

# Define a function to format the columns with fixed width


def format_cell(value, width):
    return str(value).ljust(width)


# Set fixed width for each column
col_widths = [15, 15, 10]  # Adjust these values as per your column widths

# Print the header (first row)
for column in range(1, sheet.max_column + 1):
    print(format_cell(sheet.cell(1, column).value,
          col_widths[column-1]), end=' ')
print('')

# Print the rest of the rows (starting from the second row)
for row in range(2, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row, column).value
        print(format_cell(cell_value, col_widths[column-1]), end=' ')
    print('')
